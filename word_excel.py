import os
import tkinter as tk
from tkinter import BooleanVar, filedialog, Text, Button, Checkbutton, Label
from docx import Document
from openpyxl import Workbook
import datetime

def on_select_word():
    global word_filepath
    word_filepath = filedialog.askopenfilename(defaultextension=".docx", filetypes=[("Word Document", "*.docx")])
    if not word_filepath.endswith('.docx'):
        word_filepath = ""
        word_label.config(text="Invalid file type selected")
        raise ValueError("Invalid file type selected. Please select a Word (.docx) file.")
    else:
        word_filepath_short = shorten_path(word_filepath)
        word_label.config(text=word_filepath_short)
    check_convert_state()

def on_select_excel():
    global excel_filepath
    excel_filepath = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel Document", "*.xlsx")])
    if not excel_filepath.endswith('.xlsx'):
        excel_filepath = ""
        excel_label.config(text="Invalid file type selected")
        raise ValueError("Invalid file type selected. Please select an Excel (.xlsx) file.")
    else:
        excel_filepath_short = shorten_path(excel_filepath)
        excel_label.config(text=excel_filepath_short)
    check_convert_state()


def on_create_excel():
    global excel_filepath
    global create_new_excel
    create_new_excel = True
    excel_filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Document", "*.xlsx")])
    excel_filepath_short = shorten_path(excel_filepath)
    excel_label.config(text=excel_filepath_short)
    check_convert_state()

def check_convert_state():
    global word_filepath
    global excel_filepath
    global create_new_excel
    log_text.delete(1.0,tk.END)
    if word_filepath != "" and (excel_filepath != "" or create_new_excel):
        convert_button.config(state='normal')
    else:
        convert_button.config(state='disabled')


def shorten_path(path):
    head, tail = os.path.split(path)
    return tail


def on_convert():
    global word_filepath
    global excel_filepath
    global create_new_excel
    log_text.config(state='normal')
    log_text.delete(1.0,tk.END)
    if create_new_excel.get():
        log_text.insert(tk.END, "Creating new Excel file...\n", 'green')
    else:
        log_text.insert(tk.END, f"Reading Excel file at {excel_filepath}...\n", 'green')
    try:
        if not os.path.exists(word_filepath):
            raise FileNotFoundError(f"Word file not found at {word_filepath}")
        if not os.path.exists(excel_filepath) and not create_new_excel.get():
            raise FileNotFoundError(f"Excel file not found at {excel_filepath}")
        doc = Document(word_filepath)
        wb = Workbook()
        ws = wb.active
        current_row = 1
        for table in doc.tables:
            for i in range(len(table.rows)):
                for j in range(len(table.columns)):
                    ws.cell(row=current_row, column=j+1).value = table.cell(i, j).text
                current_row += 1
            log_text.insert(tk.END, f"Table {table} successfully converted...\n", 'green')
        wb.save(excel_filepath)
        log_text.insert(tk.END, f"All tables successfully converted and saved to {excel_filepath}.\n", 'green')
    except FileNotFoundError as e:
        log_text.insert(tk.END, f"{e}\n", 'red')
    except Exception as e:
        log_text.insert(tk.END, f"An error occurred during the conversion: {e}\n", 'red')
    time_log = datetime.datetime.now()
    time_log = time_log.strftime("%H:%M:%S")
    log_text.insert(tk.END, f"\n\n-------- {time_log} --------")




root = tk.Tk()
root.geometry("700x500")
root.title("Word to Excel Converter")
root.resizable(False, False) #set the resizing to false
word_filepath = ""
excel_filepath = ""
create_new_excel = BooleanVar()
create_new_excel.set(False)




word_label = Label(root, text="No file selected", font=("Arial", 14))
word_label.grid(row=0, column=0, padx=5, pady=5, sticky='W')

excel_label = Label(root, text="No file selected", font=("Arial", 14))
excel_label.grid(row=1, column=0, padx=5, pady=5, sticky='W')


create_excel_button = Button(root, text="Create Excel file", command=on_create_excel, font=("Arial", 14))
create_excel_button.grid(row=1, column=2, padx=5, pady=5,sticky='E')

create_excel_check = Checkbutton(root, text="Create new file", variable=create_new_excel, state='disabled', font=("Arial", 14))
create_excel_check.grid(row=1, column=3, padx=5, pady=5,sticky='E')

convert_button = Button(root, text="Convert", command=on_convert, state='disabled', font=("Arial", 14))
convert_button.grid(row=2, column=1, padx=5, pady=5,sticky = 'NSEW')

log_text = Text(root, state='disabled', height=20, width=80, font=("Arial", 14))
log_text.grid(row=3, column=0, columnspan=4, padx=5, pady=5,sticky = 'NSEW')

log_text.tag_config('green', foreground='green')
log_text.tag_config('red', foreground='red')
log_text.tag_config('yellow', foreground='yellow')

try:
    select_word_button = Button(root, text="Select Word file", command=on_select_word, font=("Arial", 14))
    select_word_button.grid(row=0, column=1, padx=5, pady=5,sticky='E')

    select_excel_button = Button(root, text="Select Excel file", command=on_select_excel, font=("Arial", 14))
    select_excel_button.grid(row=1, column=1, padx=5, pady=5,sticky='E')
except ValueError as e:
    log_text.insert(tk.END, f"{e}\n", 'red')


root.mainloop()